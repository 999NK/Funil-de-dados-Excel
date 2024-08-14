import flet as ft
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import time

def process_excel(input_file, output_file, status_text, page):
    try:
        start_time = time.time()
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Arquivo não encontrado: {input_file}")

        def remove_decimal_part(s):
            if isinstance(s, str) and '.0' in s:
                return s.replace('.0', '')
            return s

        def format_row(row):
            formatted_row = row[2:-1]
            return formatted_row

        status_text.value = "Carregando o arquivo de entrada..."
        status_text.color = "blue"
        page.update()

        wb = load_workbook(input_file)
        last_sheet_name = wb.sheetnames[-1]
        data = pd.read_excel(input_file, sheet_name=last_sheet_name)

        status_text.value = f"Arquivo carregado com sucesso. Usando a última planilha: {last_sheet_name}."
        status_text.color = "blue"
        page.update()

        if os.path.exists(output_file):
            out_wb = load_workbook(output_file)
        else:
            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active)

        chunk_size = 1000
        modified_sheets = []  # Lista para rastrear as planilhas modificadas

        status_text.value = "Iniciando processamento dos dados..."
        status_text.color = "blue"
        page.update()

        for start in range(0, len(data), chunk_size):
            chunk = data.iloc[start:start + chunk_size].copy()
            chunk['concat'] = chunk.iloc[:, 0].astype(str) + '-' + chunk.iloc[:, 1].astype(str)
            chunk['concat'] = chunk['concat'].apply(remove_decimal_part)
            filtered_chunk = chunk[chunk['concat'].apply(lambda x: x.split('-')[0] in chunk.iloc[:, 0].astype(str).values)]
            grouped_data = filtered_chunk.groupby('concat')

            styles = {
                'currency': 'R$ #,##0.00',
                'percentage': '0.00%',
                'date': 'DD/MM/YYYY',
                'general': 'General'
            }
            font_style = Font(size=9)

            for sheet_name, group_df in grouped_data:
                if sheet_name in out_wb.sheetnames:
                    sheet = out_wb[sheet_name]
                else:
                    sheet = out_wb.create_sheet(title=sheet_name)
                    headers = ['Campo Adicional', 'Assembleia', 'Valor', 'Fundo Comum', '%', 'Adm. Antecipada', 'Taxa Adm.', 'Fundo Reserva', 'Total', 'Desembolso', 'Vencimento']
                    sheet.append(headers)
                    modified_sheets.append(sheet_name)  # Adiciona o nome da planilha à lista

                last_row = sheet.max_row

                for row_index, row in enumerate(group_df.itertuples(index=False, name=None), start=last_row + 1):
                    formatted_row = format_row(row)
                    if last_row > 1:
                        previous_value = sheet.cell(row=last_row, column=1).value
                        new_value = previous_value + 1
                    else:
                        new_value = 1
                    sheet.append([new_value] + list(formatted_row))
                    last_row = sheet.max_row

                for row in sheet.iter_rows(min_row=last_row - len(group_df) + 1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.number_format = styles['general'] if cell.column == 1 else styles['currency']
                        cell.font = font_style

                    row[0].number_format = styles['general']
                    row[1].number_format = styles['currency']
                    row[2].number_format = styles['currency']
                    row[3].number_format = styles['percentage']
                    row[4].number_format = styles['general']
                    row[5].number_format = styles['currency']
                    row[6].number_format = styles['percentage']
                    row[7].number_format = styles['currency']
                    row[8].number_format = styles['percentage']
                    row[9].number_format = styles['currency']
                    row[10].number_format = styles['general']
                    row[11].number_format = styles['date']

        status_text.value = "Salvando os dados no arquivo de saída..."
        status_text.color = "blue"
        page.update()
        
        out_wb.save(output_file)
        process_time = time.time() - start_time
        status_text.value = f"Processamento concluído em {process_time:.2f} segundos."
        status_text.color = "green"
        page.update()

    except FileNotFoundError:
        status_text.value = f"Arquivo não encontrado: {input_file}"
        status_text.color = "red"
        page.update()
    except Exception as e:
        status_text.value = f"Ocorreu um erro: {e}"
        status_text.color = "red"
        page.update()

def main(page: ft.Page):
    page.title = "Processador de Arquivos Excel"
    page.theme_mode = ft.ThemeMode.DARK
    page.window_width = 500
    page.window_height = 600
    page.vertical_alignment = ft.MainAxisAlignment.START  # Alinhamento vertical para o topo
    page.horizontal_alignment = ft.MainAxisAlignment.CENTER  # Alinhamento horizontal centralizado

    input_file = None
    output_file = None

    def file_picker_result(e):
        nonlocal input_file, output_file
        if e.control == input_file_picker:
            if e.files:
                input_file = e.files[0].path
                status_text.value = f"Arquivo de entrada selecionado: {input_file}"
                status_text.color = "blue"
                page.update()
        elif e.control == output_file_picker:
            if e.files:
                output_file = e.files[0].path
                status_text.value = f"Arquivo de saída selecionado: {output_file}"
                status_text.color = "blue"
                page.update()

    def start_processing(e):
        if input_file and output_file:
            process_excel(input_file, output_file, status_text, page)
        else:
            status_text.value = "Por favor, selecione os arquivos de entrada e saída."
            status_text.color = "red"
            page.update()

    # Criação dos controlos
    input_file_picker = ft.FilePicker(on_result=file_picker_result)
    output_file_picker = ft.FilePicker(on_result=file_picker_result)
    status_text = ft.Text(value="", color="black", size=16)

    select_input_button = ft.ElevatedButton("Selecionar Arquivo de Entrada", on_click=lambda _: input_file_picker.pick_files())
    select_output_button = ft.ElevatedButton("Selecionar Arquivo de Saída", on_click=lambda _: output_file_picker.pick_files())
    start_button = ft.ElevatedButton("Iniciar Processamento", on_click=start_processing)

    # Adicionando os controlos à página
    page.add(input_file_picker)
    page.add(output_file_picker)
    
    # Adicionando os botões e status à página
    page.add(
        ft.Column(
            controls=[
                ft.Container(height=15),  # Espaço em branco de 15px acima do primeiro botão
                select_input_button,
                select_output_button,
                start_button,
                ft.Container(height=2, bgcolor="blue"),  # Linha azul separadora
                status_text
            ],
            alignment=ft.MainAxisAlignment.START,  # Alinhamento vertical para o topo
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,  # Alinhamento horizontal centralizado
            expand=True  # Expande a coluna para preencher o espaço disponível
        )
    )

# Definir o app
ft.app(target=main)
